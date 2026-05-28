import urllib.request
import re
import ssl
from datetime import datetime, timedelta
import time
import email.utils
from urllib.parse import urlparse, urljoin
import os
import openpyxl
# ★変更点：Alignment（配置設定）を追加インポート
from openpyxl.styles import Font, Alignment

# ==========================================
# === URL・フォルダ設定エリア ===
# ==========================================
base_folder = r'\\192.168.2.241\share\disk1\GATEWING MA\顧客情報\00_マニアプロデュース様_お知らせ'
_save_folder = base_folder
_filename    = datetime.now().strftime('%Y年%m月') + '_お知らせ内容リスト(2年間).xlsx'
output_file  = os.path.join(_save_folder, _filename)

# マニアシリーズ 6店舗
TARGET_SITES = [
    {'店舗名': '小籠包マニア 中目黒本店', 'URL': 'https://xiaolongbaomania-nakameguro.com/'},
    {'店舗名': '焼き小籠包マニア', 'URL': 'https://yaki-xiaolongbaomania.com/'},
    {'店舗名': '北京ダックマニア 虎ノ門ヒルズ本店', 'URL': 'https://pekingduckmania-toranomon.com/'},
    {'店舗名': '餃子マニア 虎ノ門ヒルズ店', 'URL': 'https://gyozamania-toranomon.com/'},
    {'店舗名': '餃子マニア 品川別館', 'URL': 'https://gyozamania-bekkan.com/'},
    {'店舗名': '餃子マニア 品川本店', 'URL': 'https://gyozamania-honten.com/'},
]

# 直近2年（730日前）を基準日とする
border_date = datetime.now() - timedelta(days=730)
print(f"基準日: {border_date.strftime('%Y年%m月%d日')} 以降の、直近2年分の記事内容を抽出します\n")

# ==========================================
# === システム処理エリア（変更不要） ===
# ==========================================

# SSL設定
ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE
HEADERS = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

# 日付パターン
DATE_PATTERN = r'(20[1-9][0-9])(?:年|\.|/|-)(0?[1-9]|1[0-2])(?:月|\.|/|-)(0?[1-9]|[12][0-9]|3[01])(?:日)?'

# お知らせ関連キーワード
NEWS_KW_JP = ['お知らせ', 'ニュース', '新着情報', '最新情報', 'トピックス', 'インフォメーション', '更新情報']
NEWS_KW_EN = ['news', 'information', 'topics', 'notice', 'announce', 'info', 'update', 'blog']
NEWS_PATH  = ['news', 'information', 'topics', 'info', 'notice', 'announce', 'blog', 'press', 'journal']


def fetch_html(url, timeout=15):
    """HTMLを取得（Shift-JIS / EUC-JP / UTF-8 対応）"""
    url = re.sub(r'^https?://https?://', 'https://', url)
    req = urllib.request.Request(url, headers=HEADERS)
    try:
        response = urllib.request.urlopen(req, context=ctx, timeout=timeout)
        raw = response.read()
    except Exception as e:
        raise e
        
    enc_match = re.search(rb'charset=["\']?\s*([a-zA-Z0-9\-_]+)', raw[:4096])
    detected = enc_match.group(1).decode('ascii', errors='ignore').lower() if enc_match else 'utf-8'
    detected = detected.replace('shift-jis', 'shift_jis').replace('x-sjis', 'shift_jis')
    for enc in [detected, 'utf-8', 'shift_jis', 'euc-jp', 'iso-2022-jp']:
        try:
            return raw.decode(enc)
        except Exception:
            pass
    return raw.decode('utf-8', errors='ignore')

def to_date(year, month, day):
    try:
        y, m, d = int(year), int(month), int(day)
        if 2010 <= y <= datetime.now().year:
            return datetime(y, m, d)
    except Exception:
        pass
    return None

def strip_tags(html):
    """HTMLタグを除去してプレーンテキストに変換"""
    if not html:
        return ""
    html = re.sub(r'<(script|style)[^>]*>.*?</\1>', ' ', html, flags=re.DOTALL | re.IGNORECASE)
    html = re.sub(r'<[^>]+>', ' ', html)
    html = html.replace('&nbsp;', ' ').replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
    html = re.sub(r'\s+', ' ', html)
    return html.strip()

def clean_sheet_title(title):
    """Excelのシート名に使えない記号を除去し、31文字以内に収める"""
    invalid_chars = r'[\\/\?\*\[\]]'
    clean = re.sub(invalid_chars, '_', title)
    return clean[:30]


# ===== 方法① RSSフィード =====
def from_rss(base_url, border_date):
    feed_candidates = [
        base_url.rstrip('/') + '/feed/',
        base_url.rstrip('/') + '/?feed=rss2',
        base_url.rstrip('/') + '/feed/rss/',
    ]
    today = datetime.now()
    articles = []

    for feed_url in feed_candidates:
        try:
            raw = fetch_html(feed_url, timeout=10)
            items = re.findall(r'<item[^>]*>(.*?)</item>', raw, re.DOTALL | re.IGNORECASE)
            
            for item in items:
                d = None
                pub = re.search(r'<pubDate>([^<]+)</pubDate>', item, re.IGNORECASE)
                if pub:
                    parsed = email.utils.parsedate(pub.group(1).strip())
                    if parsed:
                        d = to_date(parsed[0], parsed[1], parsed[2])
                
                dc = re.search(r'<dc:date>([^<]+)</dc:date>', item, re.IGNORECASE)
                if dc and not d:
                    m = re.match(r'(20\d{2})-(0?[1-9]|1[0-2])-(0?[1-9]|[12]\d|3[01])', dc.group(1))
                    if m:
                        d = to_date(m.group(1), m.group(2), m.group(3))
                
                if d and border_date <= d <= today:
                    title_match = re.search(r'<title[^>]*>(.*?)</title>', item, re.DOTALL | re.IGNORECASE)
                    title = title_match.group(1).strip() if title_match else 'タイトルなし'
                    title = re.sub(r'<!\[CDATA\[(.*?)\]\]>', r'\1', title, flags=re.DOTALL)
                    
                    content_match = re.search(r'<content:encoded[^>]*>(.*?)</content:encoded>', item, re.DOTALL | re.IGNORECASE)
                    if not content_match:
                        content_match = re.search(r'<description[^>]*>(.*?)</description>', item, re.DOTALL | re.IGNORECASE)
                    
                    content = content_match.group(1).strip() if content_match else ''
                    content = re.sub(r'<!\[CDATA\[(.*?)\]\]>', r'\1', content, flags=re.DOTALL)
                    
                    articles.append({
                        'date': d,
                        'title': strip_tags(title),
                        'content': strip_tags(content),
                        'method': f'RSS({feed_url})'
                    })
            if articles:
                return articles
        except Exception:
            pass
    return articles


# ===== 方法② テキストエリアから簡易抽出 =====
def from_plaintext_near_keyword(html, border_date):
    text = strip_tags(html)
    articles = []
    all_keywords = NEWS_KW_JP + [k.upper() for k in NEWS_KW_EN] + [k.capitalize() for k in NEWS_KW_EN]
    today = datetime.now()

    for keyword in all_keywords:
        pos = 0
        while True:
            idx = text.find(keyword, pos)
            if idx == -1:
                break
            
            chunk = text[idx: idx + 500]
            for y, mo, d in re.findall(DATE_PATTERN, chunk):
                dt = to_date(y, mo, d)
                if dt and border_date <= dt <= today:
                    if not any(a['date'] == dt for a in articles):
                        articles.append({
                            'date': dt,
                            'title': f'{keyword}周辺の情報',
                            'content': chunk.strip()[:400] + "...",
                            'method': 'テキストスキャン'
                        })
            pos = idx + 1
    return articles


# ===== 方法③ サブページ探索 =====
def find_news_subpages(html, base_url):
    base_netloc = urlparse(base_url).netloc
    candidates = []
    for href, link_html in re.findall(r'<a[^>]*href=["\']([^"\'#][^"\']*)["\'][^>]*>(.*?)</a>',
                                       html, re.IGNORECASE | re.DOTALL):
        link_text = re.sub(r'<[^>]+>', '', link_html).strip()
        text_hit = any(kw in link_text for kw in NEWS_KW_JP + [k.upper() for k in NEWS_KW_EN])
        path_hit = any(kw in href.lower() for kw in NEWS_PATH)
        if text_hit or path_hit:
            abs_url = urljoin(base_url, href)
            if urlparse(abs_url).netloc == base_netloc and abs_url != base_url:
                if abs_url not in candidates:
                    candidates.append(abs_url)
    return candidates[:3]


def get_site_articles(url, border_date):
    """対象URLから直近1年分の記事リストを一括取得する"""
    articles = from_rss(url, border_date)
    if articles:
        return articles

    try:
        html = fetch_html(url)
    except Exception as e:
        return [{'date': None, 'title': 'エラー', 'content': f'取得エラー: {str(e)[:80]}', 'method': 'エラー'}]

    articles = from_plaintext_near_keyword(html, border_date)
    if articles:
        return articles

    for sub_url in find_news_subpages(html, url):
        try:
            time.sleep(0.5)
            articles = from_rss(sub_url, border_date)
            if articles:
                return articles
            sub_html = fetch_html(sub_url)
            articles = from_plaintext_near_keyword(sub_html, border_date)
            if articles:
                short = sub_url.replace(url, '').lstrip('/')[:20]
                for a in articles:
                    a['method'] += f'[/{short}]'
                return articles
        except Exception:
            pass

    return []


# ===== メイン実行処理 =====
print(f"💡 設定されたマニアシリーズ計 {len(TARGET_SITES)} 店舗を順に処理します。\n")
print(f"{'='*60}")

# Excelの初期化
wb = openpyxl.Workbook()
is_first_sheet = True

for i, site in enumerate(TARGET_SITES, 1):
    url = site.get('URL', '').strip()
    site_name = site.get('店舗名', f'店舗_{i}')
    
    if not url or not url.startswith('http'):
        print(f"\n[{i}] {site_name} -> URLが不正なためスキップします。")
        continue

    print(f"\n[{i}] {site_name}")
    print(f"     URL: {url}")

    # 直近1年分の記事を取得
    articles = get_site_articles(url, border_date)
    
    # シートの作成
    sheet_title = clean_sheet_title(site_name)
    if is_first_sheet:
        ws = wb.active
        ws.title = sheet_title
        is_first_sheet = False
    else:
        ws = wb.create_sheet(title=sheet_title)
        
    ws.append(['投稿日', 'タイトル', '本文（抽出内容）', '取得方法'])
    
    if articles and articles[0]['title'] != 'エラー':
        articles.sort(key=lambda x: x['date'] if x['date'] else datetime.min, reverse=True)
        print(f"     → 直近1年の記事を {len(articles)} 件検出しました。")
        for a in articles:
            date_str = a['date'].strftime('%Y年%m月%d日') if a['date'] else '不明'
            
            # ★変更点①：「。」を「。\n」に置換してセル内改行を仕込む
            formatted_content = a['content'].replace('。', '。\n').strip()
            
            ws.append([date_str, a['title'], formatted_content, a['method']])
    elif articles and articles[0]['title'] == 'エラー':
        print(f"     → {articles[0]['content']}")
        ws.append(['エラー', articles[0]['content'], '', ''])
    else:
        print("     → 直近1年の記事は検出されませんでした。")
        ws.append(['該当なし', '直近1年のお知らせは見つかりませんでした。', '', ''])
        
    # ★変更点②：見栄え（フォント・固定幅・折り返し・上揃え）の大幅調整
    font = Font(name='游ゴシック')
    # 本文が見やすくなるように、各列の幅を設定（C列を広めの60に固定）
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 60 # 本文列
    ws.column_dimensions['D'].width = 25
    
    align_wrap = Alignment(wrap_text=True, vertical='top') # 折り返しON ＋ 上揃え
    align_normal = Alignment(vertical='top')              # 改行しない列も上揃えに
    
    for row_idx, excel_row in enumerate(ws.iter_rows(), start=1):
        for col_idx, cell in enumerate(excel_row, start=1):
            cell.font = font
            
            # 2行目（データ行）以降のスタイル適用
            if row_idx > 1:
                if col_idx == 3: # 本文の列（C列）だけ折り返しを有効化
                    cell.alignment = align_wrap
                else:
                    cell.alignment = align_normal

    time.sleep(1)

print(f"\n{'='*60}")
os.makedirs(_save_folder, exist_ok=True)
wb.save(output_file)
print(f"すべての処理が完了しました！\n出力ファイル: {output_file}")